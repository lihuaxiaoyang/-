import asyncio
from pyppeteer import launch
import openpyxl


async def main():
    browser = await launch({'headless': True, 'args': ['--no-sandbox']})
    page = await browser.newPage()

    await page.goto(
        'http://bmfw.www.gov.cn/yqfxdjcx/risk.html')
    # await asyncio.sleep(1)

    data_list = ["高风险区"]
    data_lists = ["高风险区"]

    # 高风险数据写入
    while (True):
        items_high = await page.xpath('//div[@class="h-container"]/div[@class="h-header"]')
        for items_for_high in items_high:
            dist = {}
            text_for_high = await (await items_for_high.getProperty('innerText')).jsonValue()
            process_text = text_for_high.split('\n')
            result_str = ""
            temp_list = process_text[0].split(' ')
            dist['0'] = temp_list[0]
            dist['1'] = temp_list[1]
            dist['2'] = temp_list[2]
            for item in temp_list:
                result_str += item
            dist['3'] = process_text[1].split('\t')[0]
            result_str += process_text[1].split('\t')[0]
            data_list.append(result_str)
            data_lists.append(dist)
        items_nextPage = await page.xpath('//div[@class="h-content"]/div[@class="bottom-text"]/div[@class="pages-box"]')
        items_nextPage_html = await (await items_nextPage[0].getProperty('innerHTML')).jsonValue()
        if ('<button type="button" id="nextPage" disabled="">下一页</button>' in items_nextPage_html):
            break
        else:
            await page.click('.h-content button:nth-last-child(2)')

    # 中风险区写入
    data_list.append("中风险区")
    data_lists.append("中风险区")
    await page.click('.r-middle')
    while (True):
        items_high = await page.xpath('//div[@class="m-container"]/div[@class="m-header"]')
        for items_for_high in items_high:
            dist = {}
            text_for_high = await (await items_for_high.getProperty('innerText')).jsonValue()
            process_text = text_for_high.split('\n')
            result_str = ""
            temp_list = process_text[0].split(' ')
            dist['0'] = temp_list[0]
            dist['1'] = temp_list[1]
            dist['2'] = temp_list[2]
            for item in temp_list:
                result_str += item
            dist['3'] = process_text[1].split('\t')[0]
            result_str += process_text[1].split('\t')[0]
            data_list.append(result_str)
            data_lists.append(dist)
        items_nextPage = await page.xpath('//div[@class="m-content"]/div[@class="bottom-text"]/div[@class="pages-box"]')
        items_nextPage_html = await (await items_nextPage[0].getProperty('innerHTML')).jsonValue()
        if ('<button type="button" id="nextPage" disabled="">下一页</button>' in items_nextPage_html):
            break
        else:
            await page.click('.m-content button:nth-last-child(2)')

    # 保存为xls格式
    path = r"C:\Users\16397\Desktop\中高风险地区疫情数据.xls"
    sheetStr = 'sheet1'
    # write_to_excel(path, sheetStr, data_list)
    write_to_excels(path, sheetStr, data_lists)

    await browser.close()


async def wait_fornavigation(page, events):  # 等到某动作完成
    await asyncio.wait([
        events,
        page.waitForNavigation({'timeout': 50000}),
    ])


def write_to_excel(path, sheetStr, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheetStr

    count = 0

    for i in range(len(data)):
        sheet.cell(row=i + 1, column=1, value=data[i])
        if (data[i].find("郑州") != -1):
            count += 1
            print(data[i])
    print("中高风险区总计：{}个".format(count))

    workbook.save(path)


def write_to_excels(path, sheetStr, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheetStr

    for i in range(len(data)):
        # print(data[i])
        if (data[i] == "高风险区" or data[i] == "中风险区"):
            sheet.cell(row=i+1,column=1,value=data[i])
        else:
            for j in range(len(data[i])):
                sheet.cell(row=i + 1, column=j + 1, value=data[i][str(j)])
    workbook.save(path)


# <button type="button" id="nextPage" disabled="">下一页</button>

if __name__ == '__main__':
    asyncio.get_event_loop().run_until_complete(main())
